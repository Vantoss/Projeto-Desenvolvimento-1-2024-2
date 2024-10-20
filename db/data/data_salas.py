
import pandas as pd
import openpyxl

dataframe1 = pd.read_excel("reservas.xlsx")

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("reservas.xlsx")

# Define variable to read sheet
dataframe1 = dataframe.active

# Iterate the loop to read the cell values
f = open("salas.txt","w", encoding="utf-8")

# f.write("(")
count = 0
for col in dataframe1.iter_cols(1, dataframe1.max_column):

    if col[1].value != None:
               
        f.write("(" + str(col[1].value) + ", " + str(col[0].value)+ ")\n")
        
f.close()

