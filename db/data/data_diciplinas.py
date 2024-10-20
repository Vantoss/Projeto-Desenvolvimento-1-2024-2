import pandas as pd
import openpyxl

sheets = ["janeiro.xlsx",
          "fevereiro.xlsx",
          "marco.xlsx",
          "abril.xlsx",
          "maio.xlsx",
          "junho.xlsx",
          "julho.xlsx",
          "agosto.xlsx",
          "setembro.xlsx",
          "outubro.xlsx",
          "novembro.xlsx",
          "dezembro.xlsx",
          ]

diciplinas = set()

for sheet in sheets:

    dataframe1 = pd.read_excel("sheets/" + sheet)

# Define variable to load the dataframe
    dataframe = openpyxl.load_workbook("sheets/" + sheet) 

# Define variable to read sheet
    dataframe1 = dataframe.active

# Iterate the loop to read the cell values

    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        
        if col[2].value == "Turma":
            for row in range(3, dataframe1.max_row):
                if col[row].value:
                    s = col[row].value
                    diciplinas.add(s)
            

f = open("diciplinas.txt","w",encoding="utf-8")


for diciplina in diciplinas:
    f.write(diciplina)
    f.write("\n")

f.close()